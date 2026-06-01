"""
Admin Licenses API — to'liq boshqaruv: statistika, ro'yxat, tafsilot,
yangilash, bekor qilish, yaratish, ommaviy amallar, Excel eksport.
"""
from datetime import timedelta

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import License, LicenseType


ADMIN_ROLES = {'super_admin', 'region_admin', 'staff', 'viewer'}
WRITE_ROLES = {'super_admin', 'region_admin', 'staff'}
REVOKE_ROLES = {'super_admin'}


class AdminLicensePermission(IsAuthenticated):
    """Authenticated admin (any admin role) — read access."""
    def has_permission(self, request, view):
        if not super().has_permission(request, view):
            return False
        return getattr(request.user, 'role', None) in ADMIN_ROLES


def _scope_queryset(request, qs):
    """region_admin faqat o'z viloyat litsenziyalarini ko'rsin."""
    user = request.user
    if getattr(user, 'role', None) == 'region_admin' and getattr(user, 'region_id', None):
        qs = qs.filter(Q(region_id=user.region_id) | Q(user__region_id=user.region_id))
    return qs


def _write_audit(user, action, entity_id, data):
    try:
        from apps.audit.models import AuditLog
        AuditLog.objects.create(
            user=user, action=action,
            entity_type='license', entity_id=str(entity_id),
            new_data=data if isinstance(data, dict) else {'data': str(data)},
        )
    except Exception:
        pass


def _initials(name: str) -> str:
    parts = [p for p in (name or '').strip().split() if p]
    if not parts:
        return 'XX'
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[1][0]).upper()


def _serialize_license(lic: License) -> dict:
    comp = lic.computed_status
    days_left = lic.days_until_expiry
    user = lic.user
    return {
        'id': str(lic.id),
        'license_number': lic.license_number,
        'user': {
            'id': str(user.id),
            'full_name': user.full_name or user.phone,
            'phone': user.phone,
            'email': user.email or '',
            'avatar_url': user.avatar.url if getattr(user, 'avatar', None) else None,
            'initials': _initials(user.full_name or user.phone),
        },
        'license_type': {
            'id': lic.license_type_id,
            'code': lic.license_type.code,
            'name': lic.license_type.name_uz,
            'color': lic.license_type.color_hex or '#1A56A0',
            'category': lic.license_type.category,
        },
        'region': (lic.region.name_uz if lic.region else
                   (user.region.name_uz if user.region else '')),
        'status': comp,
        'status_display': {
            'active': 'Faol',
            'expired': "Muddati o'tgan",
            'suspended': "To'xtatilgan",
            'revoked': 'Bekor qilingan',
        }.get(comp, comp),
        'issued_at': lic.issued_at.strftime('%Y-%m-%d') if lic.issued_at else '',
        'expires_at': lic.expires_at.strftime('%Y-%m-%d') if lic.expires_at else '',
        'days_left': days_left,
        'is_expiring_soon': 0 < days_left <= 30,
        'pdf_url': lic.pdf_url or '',
        'qr_code_url': lic.qr_code_url or '',
    }


# ─────────────────────────────────────────────────────
# 1. STATISTIKA
# ─────────────────────────────────────────────────────
class AdminLicenseStatsView(APIView):
    permission_classes = [AdminLicensePermission]

    def get(self, request):
        now = timezone.now()
        qs = _scope_queryset(request, License.objects.all())

        # Joriy son
        total = qs.count()
        active = qs.filter(is_active=True, status='active', expires_at__gt=now).count()
        expired = qs.filter(is_active=True, expires_at__lt=now).count()
        suspended = qs.filter(status='suspended', is_active=True).count()
        revoked = qs.filter(is_active=False).count()

        # 30 kunda muddati tugaydiganlar
        expiring_soon = qs.filter(
            is_active=True, expires_at__gt=now,
            expires_at__lte=now + timedelta(days=30)
        ).count()

        # O'tgan oy bilan solishtirish
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        prev_month_end = month_start - timedelta(seconds=1)
        prev_month_start = prev_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        def growth(curr, prev):
            if not prev:
                return 100.0 if curr else 0.0
            return round(((curr - prev) / prev) * 100, 1)

        prev_total = qs.filter(issued_at__lt=prev_month_end).count()
        prev_active = qs.filter(
            is_active=True, expires_at__gt=prev_month_end,
            issued_at__lt=prev_month_end,
        ).count()
        prev_expired = qs.filter(expires_at__lt=prev_month_end, is_active=True).count()
        prev_suspended = qs.filter(
            status='suspended', is_active=True,
            suspended_at__lt=prev_month_end,
        ).count()

        return Response({
            'total': total,
            'active': active,
            'expired': expired,
            'suspended': suspended,
            'revoked': revoked,
            'expiring_soon': expiring_soon,
            'growth': {
                'total': growth(total, prev_total),
                'active': growth(active, prev_active),
                'expired': growth(expired, prev_expired),
                'suspended': growth(suspended, prev_suspended),
            },
        })


# ─────────────────────────────────────────────────────
# 2. RO'YXAT
# ─────────────────────────────────────────────────────
class AdminLicenseListView(APIView):
    permission_classes = [AdminLicensePermission]

    def get(self, request):
        now = timezone.now()
        qs = License.objects.select_related(
            'user', 'user__region', 'license_type', 'region', 'application'
        ).order_by('-issued_at')
        qs = _scope_queryset(request, qs)

        # Filtrlar
        p = request.query_params
        search = (p.get('search') or '').strip()
        status_f = (p.get('status') or '').strip()
        type_f = (p.get('license_type') or '').strip()
        region_f = (p.get('region') or '').strip()
        date_from = (p.get('date_from') or '').strip()
        date_to = (p.get('date_to') or '').strip()
        expiring = p.get('expiring_soon') == '1'

        if search:
            qs = qs.filter(
                Q(license_number__icontains=search) |
                Q(user__full_name__icontains=search) |
                Q(user__phone__icontains=search) |
                Q(user__email__icontains=search)
            )

        if status_f == 'active':
            qs = qs.filter(is_active=True, status='active', expires_at__gt=now)
        elif status_f == 'expired':
            qs = qs.filter(is_active=True, expires_at__lt=now)
        elif status_f == 'suspended':
            qs = qs.filter(status='suspended', is_active=True)
        elif status_f == 'revoked':
            qs = qs.filter(is_active=False)

        if type_f:
            qs = qs.filter(license_type__code=type_f)
        if region_f:
            qs = qs.filter(Q(region_id=region_f) | Q(user__region_id=region_f))
        if date_from:
            qs = qs.filter(issued_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(issued_at__date__lte=date_to)
        if expiring:
            qs = qs.filter(
                is_active=True, expires_at__gt=now,
                expires_at__lte=now + timedelta(days=30),
            )

        # Pagination
        try:
            limit = max(1, min(int(p.get('limit', 20)), 100))
            offset = max(0, int(p.get('offset', 0)))
        except (TypeError, ValueError):
            limit, offset = 20, 0

        total = qs.count()
        items = list(qs[offset:offset + limit])

        return Response({
            'count': total,
            'limit': limit,
            'offset': offset,
            'results': [_serialize_license(l) for l in items],
        })


# ─────────────────────────────────────────────────────
# 3. TAFSILOT
# ─────────────────────────────────────────────────────
class AdminLicenseDetailView(APIView):
    permission_classes = [AdminLicensePermission]

    def get(self, request, pk):
        try:
            lic = License.objects.select_related(
                'user', 'user__region', 'license_type', 'region', 'application'
            ).get(pk=pk)
        except License.DoesNotExist:
            return Response({'detail': 'Topilmadi'}, status=404)

        # Scope check
        scoped = _scope_queryset(request, License.objects.filter(pk=pk)).exists()
        if not scoped:
            return Response({'detail': "Ruxsat yo'q"}, status=403)

        data = _serialize_license(lic)
        # Qo'shimcha tafsilot
        u = lic.user
        data['user'].update({
            'coaching_years': getattr(u, 'coaching_years', None),
            'job_title': u.job_title or '',
            'workplace': u.workplace or '',
            'birth_date': u.birth_date.isoformat() if u.birth_date else None,
        })
        data['revoke_reason'] = lic.revoke_reason or ''
        data['suspend_reason'] = lic.suspend_reason or ''
        data['revoked_at'] = lic.revoked_at.isoformat() if lic.revoked_at else None
        data['suspended_at'] = lic.suspended_at.isoformat() if lic.suspended_at else None
        data['application_id'] = str(lic.application_id) if lic.application_id else None
        return Response(data)


# ─────────────────────────────────────────────────────
# 4. TAHRIRLASH (muddat uzaytirish, suspend/activate)
# ─────────────────────────────────────────────────────
class AdminLicenseUpdateView(APIView):
    permission_classes = [AdminLicensePermission]

    def patch(self, request, pk):
        if request.user.role not in WRITE_ROLES:
            return Response({'detail': "Ruxsat yo'q"}, status=403)

        try:
            lic = License.objects.get(pk=pk)
        except License.DoesNotExist:
            return Response({'detail': 'Topilmadi'}, status=404)

        if not _scope_queryset(request, License.objects.filter(pk=pk)).exists():
            return Response({'detail': "Ruxsat yo'q"}, status=403)

        data = request.data

        # Muddat uzaytirish
        if 'extends_days' in data:
            try:
                days = int(data['extends_days'])
            except (TypeError, ValueError):
                return Response({'detail': 'extends_days butun son bo\'lishi kerak'}, status=400)
            if days <= 0 or days > 730:
                return Response({'detail': "Muddat 1-730 kun oralig'ida bo'lishi kerak"}, status=400)
            lic.expires_at = lic.expires_at + timedelta(days=days)

        action = data.get('action')

        if action == 'suspend':
            reason = (data.get('reason') or '').strip()
            if not reason:
                return Response({'detail': 'Sabab kiritish shart'}, status=400)
            lic.status = 'suspended'
            lic.suspend_reason = reason
            lic.suspended_at = timezone.now()

        elif action == 'activate':
            if not lic.is_active:
                return Response({'detail': 'Bekor qilingan litsenziyani faollashtirib bo\'lmaydi'}, status=400)
            lic.status = 'active'
            lic.suspend_reason = ''
            lic.suspended_at = None

        lic.save()
        _write_audit(request.user, 'license_update', lic.id, dict(data))
        return Response({'success': True, 'license': _serialize_license(lic)})


# ─────────────────────────────────────────────────────
# 5. BEKOR QILISH
# ─────────────────────────────────────────────────────
class AdminLicenseRevokeView(APIView):
    permission_classes = [AdminLicensePermission]

    def post(self, request, pk):
        if request.user.role not in REVOKE_ROLES:
            return Response({'detail': 'Faqat Super Admin bekor qila oladi'}, status=403)

        reason = (request.data.get('reason') or '').strip()
        if not reason:
            return Response({'detail': 'Bekor qilish sababi kiritish shart'}, status=400)

        try:
            lic = License.objects.get(pk=pk)
        except License.DoesNotExist:
            return Response({'detail': 'Topilmadi'}, status=404)

        if not lic.is_active:
            return Response({'detail': 'Litsenziya allaqachon bekor qilingan'}, status=400)

        lic.is_active = False
        lic.status = 'revoked'
        lic.revoke_reason = reason
        lic.revoked_at = timezone.now()
        lic.save()

        _write_audit(request.user, 'license_revoke', lic.id, {'reason': reason})
        return Response({'success': True, 'license': _serialize_license(lic)})


# ─────────────────────────────────────────────────────
# 6. YANGI YARATISH
# ─────────────────────────────────────────────────────
class AdminLicenseCreateView(APIView):
    permission_classes = [AdminLicensePermission]

    def post(self, request):
        if request.user.role not in WRITE_ROLES:
            return Response({'detail': "Ruxsat yo'q"}, status=403)

        from apps.users.models import User
        data = request.data

        user_id = data.get('user_id')
        type_code = data.get('license_type_code') or data.get('license_type')
        try:
            expires_days = int(data.get('expires_days', 365))
        except (TypeError, ValueError):
            return Response({'detail': 'expires_days butun son bo\'lishi kerak'}, status=400)
        if expires_days <= 0 or expires_days > 3650:
            return Response({'detail': "Muddat 1-3650 kun oralig'ida"}, status=400)

        try:
            user = User.objects.get(id=user_id)
        except (User.DoesNotExist, ValueError, TypeError):
            return Response({'detail': 'Foydalanuvchi topilmadi'}, status=400)

        try:
            lic_type = LicenseType.objects.get(code=type_code, is_active=True)
        except LicenseType.DoesNotExist:
            return Response({'detail': 'Litsenziya turi topilmadi'}, status=400)

        # Litsenziya raqami: UFF-{YEAR}-{CODE}-{NNNNNN}
        year = timezone.now().year
        count = License.objects.filter(
            license_type=lic_type, issued_at__year=year
        ).count() + 1
        # Conflict bo'lmasin uchun loop
        while True:
            number = f"UFF-{year}-{type_code}-{str(count).zfill(6)}"
            if not License.objects.filter(license_number=number).exists():
                break
            count += 1

        lic = License.objects.create(
            user=user,
            license_type=lic_type,
            region=user.region,
            license_number=number,
            expires_at=timezone.now() + timedelta(days=expires_days),
            is_active=True,
            status='active',
        )

        _write_audit(request.user, 'license_create', lic.id, dict(data))
        return Response({
            'success': True,
            'license_number': lic.license_number,
            'id': str(lic.id),
            'license': _serialize_license(lic),
        }, status=201)


# ─────────────────────────────────────────────────────
# 7. OMMAVIY AMALLAR
# ─────────────────────────────────────────────────────
class AdminLicenseBulkView(APIView):
    permission_classes = [AdminLicensePermission]

    def post(self, request):
        if request.user.role not in WRITE_ROLES:
            return Response({'detail': "Ruxsat yo'q"}, status=403)

        data = request.data
        ids = data.get('ids') or []
        action = data.get('action')
        reason = (data.get('reason') or '').strip()

        if not isinstance(ids, list) or not ids or not action:
            return Response({'detail': 'ids (ro\'yxat) va action talab qilinadi'}, status=400)

        qs = _scope_queryset(request, License.objects.filter(id__in=ids))
        affected = qs.count()
        if not affected:
            return Response({'detail': 'Tegishli litsenziyalar topilmadi'}, status=404)

        now = timezone.now()
        if action == 'revoke':
            if request.user.role not in REVOKE_ROLES:
                return Response({'detail': 'Faqat Super Admin bekor qila oladi'}, status=403)
            if not reason:
                return Response({'detail': 'Sabab kiritish shart'}, status=400)
            qs.update(is_active=False, status='revoked', revoke_reason=reason, revoked_at=now)

        elif action == 'suspend':
            if not reason:
                return Response({'detail': 'Sabab kiritish shart'}, status=400)
            qs.update(status='suspended', suspend_reason=reason, suspended_at=now)

        elif action == 'activate':
            qs.filter(is_active=True).update(
                status='active', suspend_reason='', suspended_at=None
            )

        elif action == 'extend':
            try:
                days = int(data.get('days', 365))
            except (TypeError, ValueError):
                return Response({'detail': 'days butun son'}, status=400)
            if days <= 0 or days > 730:
                return Response({'detail': "1-730 kun oralig'ida"}, status=400)
            for lic in qs:
                lic.expires_at = lic.expires_at + timedelta(days=days)
                lic.save(update_fields=['expires_at', 'updated_at'])

        else:
            return Response({'detail': "Noto'g'ri action"}, status=400)

        _write_audit(request.user, f'license_bulk_{action}', '-', {
            'ids': [str(i) for i in ids], 'reason': reason,
        })
        return Response({'success': True, 'affected': affected})


# ─────────────────────────────────────────────────────
# 8. EXCEL EKSPORT
# ─────────────────────────────────────────────────────
class AdminLicenseExportView(APIView):
    permission_classes = [AdminLicensePermission]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response({'detail': 'openpyxl o\'rnatilmagan'}, status=500)

        now = timezone.now()
        qs = License.objects.select_related(
            'user', 'user__region', 'license_type', 'region'
        ).order_by('-issued_at')
        qs = _scope_queryset(request, qs)

        p = request.query_params
        status_f = (p.get('status') or '').strip()
        type_f = (p.get('license_type') or '').strip()

        if status_f == 'active':
            qs = qs.filter(is_active=True, status='active', expires_at__gt=now)
        elif status_f == 'expired':
            qs = qs.filter(is_active=True, expires_at__lt=now)
        elif status_f == 'suspended':
            qs = qs.filter(status='suspended', is_active=True)
        elif status_f == 'revoked':
            qs = qs.filter(is_active=False)
        if type_f:
            qs = qs.filter(license_type__code=type_f)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Litsenziyalar"

        hdr_fill = PatternFill(start_color="0D3B6E", end_color="0D3B6E", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        center = Alignment(horizontal='center', vertical='center')
        thin = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0'),
        )

        headers = [
            'Litsenziya raqami', 'Egasi', 'Telefon', 'Litsenziya turi',
            'Hudud', 'Holat', 'Berilgan sana', 'Tugash sanasi', 'Qolgan kunlar'
        ]
        widths = [22, 30, 18, 22, 20, 18, 16, 16, 14]
        for col, (hdr, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

        STATUS_UZ = {
            'active': 'Faol', 'expired': "Muddati o'tgan",
            'suspended': "To'xtatilgan", 'revoked': 'Bekor qilingan',
        }
        STATUS_BG = {
            'active': 'E8F5E9', 'expired': 'FFEBEE',
            'suspended': 'FFF8E1', 'revoked': 'FAFAFA',
        }

        for row_num, lic in enumerate(qs, 2):
            comp = lic.computed_status
            days = lic.days_until_expiry
            region = lic.region.name_uz if lic.region else (
                lic.user.region.name_uz if lic.user.region else ''
            )
            row_data = [
                lic.license_number,
                lic.user.full_name or '',
                lic.user.phone,
                lic.license_type.name_uz,
                region,
                STATUS_UZ.get(comp, comp),
                lic.issued_at.strftime('%d.%m.%Y') if lic.issued_at else '',
                lic.expires_at.strftime('%d.%m.%Y') if lic.expires_at else '',
                max(days, 0),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin
                cell.alignment = Alignment(vertical='center')

            status_cell = ws.cell(row=row_num, column=6)
            status_cell.fill = PatternFill(
                start_color=STATUS_BG.get(comp, 'FFFFFF'),
                end_color=STATUS_BG.get(comp, 'FFFFFF'),
                fill_type="solid",
            )
            if 0 < days <= 30:
                ws.cell(row=row_num, column=9).font = Font(color="E65100", bold=True)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="uff_litsenziyalar_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response
