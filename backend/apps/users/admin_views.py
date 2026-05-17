"""
Admin Users API — to'liq boshqaruv: statistika, ro'yxat, tafsilot,
yaratish, yangilash, faollashtirish/o'chirish, Excel eksport.
"""
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Region

User = get_user_model()

ADMIN_ROLES = {'super_admin', 'region_admin', 'staff', 'viewer'}
WRITE_ROLES = {'super_admin', 'region_admin', 'staff'}
DELETE_ROLES = {'super_admin'}


class AdminUserPermission(IsAuthenticated):
    def has_permission(self, request, view):
        if not super().has_permission(request, view):
            return False
        return getattr(request.user, 'role', None) in ADMIN_ROLES


def _scope_queryset(request, qs):
    """region_admin faqat o'z viloyatidagi foydalanuvchilarni ko'rsin."""
    user = request.user
    if getattr(user, 'role', None) == 'region_admin' and getattr(user, 'region_id', None):
        qs = qs.filter(region_id=user.region_id)
    return qs


def _write_audit(user, action, entity_id, data):
    try:
        from apps.audit.models import AuditLog
        AuditLog.objects.create(
            user=user, action=action,
            entity_type='user', entity_id=str(entity_id),
            new_data=data if isinstance(data, dict) else {'data': str(data)},
        )
    except Exception:
        pass


def _initials(name: str) -> str:
    if not name:
        return '?'
    parts = [p for p in name.split() if p]
    return (''.join(p[0] for p in parts[:2]) or '?').upper()


def _serialize_user(u, request=None, with_stats=True):
    avatar_url = ''
    if u.avatar and hasattr(u.avatar, 'url'):
        try:
            avatar_url = request.build_absolute_uri(u.avatar.url) if request else u.avatar.url
        except Exception:
            avatar_url = u.avatar_url or ''
    elif u.avatar_url:
        avatar_url = u.avatar_url

    data = {
        'id': str(u.id),
        'phone': u.phone,
        'email': u.email or '',
        'full_name': u.full_name or '',
        'initials': _initials(u.full_name or u.phone),
        'first_name': u.first_name or '',
        'last_name': u.last_name or '',
        'middle_name': u.middle_name or '',
        'birth_date': u.birth_date.isoformat() if u.birth_date else None,
        'gender': u.gender or '',
        'workplace': u.workplace or '',
        'job_title': u.job_title or '',
        'coaching_years': u.coaching_years or 0,
        'role': u.role,
        'role_display': dict(User.ROLE_CHOICES).get(u.role, u.role),
        'region_id': u.region_id,
        'region': u.region.name_uz if u.region else '',
        'region_ru': u.region.name_ru if u.region else '',
        'is_active': u.is_active,
        'is_onboarded': u.is_onboarded,
        'two_factor_enabled': u.two_factor_enabled,
        'avatar_url': avatar_url,
        'language': u.language,
        'created_at': u.created_at.isoformat() if u.created_at else None,
        'last_login': u.last_login.isoformat() if u.last_login else None,
    }
    if with_stats:
        # Lightweight counts
        try:
            from apps.applications.models import Application
            from apps.licenses.models import License
            data['applications_count'] = Application.objects.filter(user=u).count()
            data['licenses_count'] = License.objects.filter(user=u).count()
            data['active_licenses_count'] = License.objects.filter(
                user=u, is_active=True, expires_at__gt=timezone.now()
            ).count()
        except Exception:
            data['applications_count'] = 0
            data['licenses_count'] = 0
            data['active_licenses_count'] = 0
    return data


# ════════════════════════════════════════════════════════════════════
# STATS
# ════════════════════════════════════════════════════════════════════
class AdminUserStatsView(APIView):
    permission_classes = [AdminUserPermission]

    def get(self, request):
        now = timezone.now()
        last_month_start = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        active_threshold = now - timedelta(hours=24)

        qs = _scope_queryset(request, User.objects.filter(deleted_at__isnull=True))

        total = qs.count()
        active = qs.filter(is_active=True).count()
        inactive = qs.filter(is_active=False).count()
        coaches = qs.filter(role='coach').count()
        admins = qs.filter(role__in=['super_admin', 'region_admin', 'staff']).count()
        active_now = qs.filter(last_login__gte=active_threshold).count()
        new_this_month = qs.filter(created_at__gte=this_month_start).count()
        new_last_month = qs.filter(
            created_at__gte=last_month_start, created_at__lt=this_month_start
        ).count()

        def pct(curr, prev):
            if not prev:
                return 100 if curr else 0
            return round((curr - prev) / prev * 100, 1)

        # By role breakdown
        by_role = list(
            qs.values('role').annotate(count=Count('id')).order_by('-count')
        )

        # By region breakdown (top 10)
        by_region = list(
            qs.filter(region__isnull=False)
              .values('region__id', 'region__name_uz', 'region__name_ru')
              .annotate(count=Count('id'))
              .order_by('-count')[:10]
        )

        return Response({
            'total': total,
            'active': active,
            'inactive': inactive,
            'coaches': coaches,
            'admins': admins,
            'active_now': active_now,
            'new_this_month': new_this_month,
            'growth': {
                'total': pct(new_this_month, new_last_month),
                'coaches': pct(
                    qs.filter(role='coach', created_at__gte=this_month_start).count(),
                    qs.filter(role='coach', created_at__gte=last_month_start, created_at__lt=this_month_start).count(),
                ),
                'admins': pct(
                    qs.filter(role__in=['super_admin', 'region_admin', 'staff'], created_at__gte=this_month_start).count(),
                    qs.filter(role__in=['super_admin', 'region_admin', 'staff'], created_at__gte=last_month_start, created_at__lt=this_month_start).count(),
                ),
            },
            'by_role': by_role,
            'by_region': by_region,
        })


# ════════════════════════════════════════════════════════════════════
# LIST
# ════════════════════════════════════════════════════════════════════
class AdminUserListView(APIView):
    permission_classes = [AdminUserPermission]

    def get(self, request):
        qs = _scope_queryset(
            request,
            User.objects.filter(deleted_at__isnull=True).select_related('region')
        )

        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(
                Q(phone__icontains=search) |
                Q(email__icontains=search) |
                Q(full_name__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(workplace__icontains=search)
            )

        role = (request.query_params.get('role') or '').strip()
        if role:
            qs = qs.filter(role=role)

        is_active = request.query_params.get('is_active')
        if is_active in ('true', '1', 'yes'):
            qs = qs.filter(is_active=True)
        elif is_active in ('false', '0', 'no'):
            qs = qs.filter(is_active=False)

        region_id = request.query_params.get('region')
        if region_id:
            try:
                qs = qs.filter(region_id=int(region_id))
            except ValueError:
                pass

        ordering = request.query_params.get('ordering') or '-created_at'
        try:
            qs = qs.order_by(ordering)
        except Exception:
            qs = qs.order_by('-created_at')

        try:
            limit = max(1, min(int(request.query_params.get('limit', 20)), 200))
            offset = max(0, int(request.query_params.get('offset', 0)))
        except ValueError:
            limit, offset = 20, 0

        count = qs.count()
        page = qs[offset:offset + limit]
        results = [_serialize_user(u, request=request, with_stats=True) for u in page]
        return Response({'count': count, 'limit': limit, 'offset': offset, 'results': results})


# ════════════════════════════════════════════════════════════════════
# DETAIL
# ════════════════════════════════════════════════════════════════════
class AdminUserDetailView(APIView):
    permission_classes = [AdminUserPermission]

    def get(self, request, user_id):
        try:
            u = _scope_queryset(request, User.objects.filter(deleted_at__isnull=True)).get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'Foydalanuvchi topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        data = _serialize_user(u, request=request, with_stats=True)

        # Recent applications
        try:
            from apps.applications.models import Application
            apps_qs = Application.objects.filter(user=u).select_related('license_type').order_by('-submitted_at')[:10]
            data['recent_applications'] = [
                {
                    'id': str(a.id),
                    'license_type_code': a.license_type.code if a.license_type else '',
                    'license_type_name': a.license_type.name_uz if a.license_type else '',
                    'status': a.status,
                    'submitted_at': a.submitted_at.isoformat() if a.submitted_at else None,
                }
                for a in apps_qs
            ]
        except Exception:
            data['recent_applications'] = []

        # Licenses
        try:
            from apps.licenses.models import License
            lic_qs = License.objects.filter(user=u).select_related('license_type').order_by('-issued_at')[:10]
            data['licenses'] = [
                {
                    'id': str(l.id),
                    'license_number': l.license_number,
                    'license_type_code': l.license_type.code if l.license_type else '',
                    'license_type_name': l.license_type.name_uz if l.license_type else '',
                    'status': l.computed_status,
                    'issued_at': l.issued_at.strftime('%Y-%m-%d') if l.issued_at else '',
                    'expires_at': l.expires_at.strftime('%Y-%m-%d') if l.expires_at else '',
                }
                for l in lic_qs
            ]
        except Exception:
            data['licenses'] = []

        return Response(data)


# ════════════════════════════════════════════════════════════════════
# UPDATE
# ════════════════════════════════════════════════════════════════════
class AdminUserUpdateView(APIView):
    permission_classes = [AdminUserPermission]

    def patch(self, request, user_id):
        if request.user.role not in WRITE_ROLES:
            return Response({'error': 'Ruxsat yo\'q'}, status=status.HTTP_403_FORBIDDEN)

        try:
            u = _scope_queryset(request, User.objects.filter(deleted_at__isnull=True)).get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'Foydalanuvchi topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        # Faqat super_admin rolni o'zgartira oladi
        if 'role' in request.data and request.user.role != 'super_admin':
            return Response({'error': 'Rolni faqat super admin o\'zgartira oladi'}, status=status.HTTP_403_FORBIDDEN)

        allowed = {
            'full_name', 'first_name', 'last_name', 'middle_name', 'email',
            'birth_date', 'gender', 'workplace', 'job_title', 'coaching_years',
            'role', 'region_id', 'is_active',
        }
        changed = {}
        for field in allowed:
            if field in request.data:
                value = request.data.get(field)
                if field == 'region_id':
                    if value in (None, '', 0):
                        u.region = None
                    else:
                        try:
                            u.region = Region.objects.get(id=int(value))
                        except (Region.DoesNotExist, ValueError):
                            return Response({'error': 'Viloyat topilmadi'}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    setattr(u, field, value)
                changed[field] = value

        # Role o'zgargan bo'lsa, is_staff'ni ham mos qilamiz
        if 'role' in changed:
            u.is_staff = u.role in {'super_admin', 'region_admin', 'staff'}

        u.save()
        _write_audit(request.user, 'update_user', u.id, changed)
        return Response(_serialize_user(u, request=request))


# ════════════════════════════════════════════════════════════════════
# ACTIVATE / DEACTIVATE
# ════════════════════════════════════════════════════════════════════
class AdminUserToggleActiveView(APIView):
    permission_classes = [AdminUserPermission]

    def post(self, request, user_id):
        if request.user.role not in WRITE_ROLES:
            return Response({'error': 'Ruxsat yo\'q'}, status=status.HTTP_403_FORBIDDEN)
        try:
            u = _scope_queryset(request, User.objects.filter(deleted_at__isnull=True)).get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'Foydalanuvchi topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        if u.id == request.user.id:
            return Response({'error': 'O\'zingizni o\'chira olmaysiz'}, status=status.HTTP_400_BAD_REQUEST)

        u.is_active = not u.is_active
        u.save(update_fields=['is_active', 'updated_at'])
        _write_audit(request.user, 'toggle_active_user', u.id, {'is_active': u.is_active})
        return Response({'id': str(u.id), 'is_active': u.is_active})


# ════════════════════════════════════════════════════════════════════
# DELETE (soft)
# ════════════════════════════════════════════════════════════════════
class AdminUserDeleteView(APIView):
    permission_classes = [AdminUserPermission]

    def delete(self, request, user_id):
        if request.user.role not in DELETE_ROLES:
            return Response({'error': 'Faqat super admin o\'chira oladi'}, status=status.HTTP_403_FORBIDDEN)
        try:
            u = User.objects.get(id=user_id, deleted_at__isnull=True)
        except User.DoesNotExist:
            return Response({'error': 'Foydalanuvchi topilmadi'}, status=status.HTTP_404_NOT_FOUND)
        if u.id == request.user.id:
            return Response({'error': 'O\'zingizni o\'chira olmaysiz'}, status=status.HTTP_400_BAD_REQUEST)
        u.deleted_at = timezone.now()
        u.is_active = False
        u.save(update_fields=['deleted_at', 'is_active', 'updated_at'])
        _write_audit(request.user, 'delete_user', u.id, {})
        return Response({'id': str(u.id), 'deleted': True})


# ════════════════════════════════════════════════════════════════════
# CREATE
# ════════════════════════════════════════════════════════════════════
class AdminUserCreateView(APIView):
    permission_classes = [AdminUserPermission]

    def post(self, request):
        if request.user.role not in WRITE_ROLES:
            return Response({'error': 'Ruxsat yo\'q'}, status=status.HTTP_403_FORBIDDEN)

        phone = (request.data.get('phone') or '').strip()
        if not phone:
            return Response({'error': 'Telefon raqam majburiy'}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(phone=phone).exists():
            return Response({'error': 'Bu telefon raqam allaqachon mavjud'}, status=status.HTTP_400_BAD_REQUEST)

        role = request.data.get('role') or 'coach'
        valid_roles = {r[0] for r in User.ROLE_CHOICES}
        if role not in valid_roles:
            return Response({'error': 'Noto\'g\'ri rol'}, status=status.HTTP_400_BAD_REQUEST)
        # region_admin staff yarata olishi mumkin, lekin admin emas
        if role in {'super_admin', 'region_admin'} and request.user.role != 'super_admin':
            return Response({'error': 'Admin yarata olmaysiz'}, status=status.HTTP_403_FORBIDDEN)

        region = None
        region_id = request.data.get('region_id')
        if region_id:
            try:
                region = Region.objects.get(id=int(region_id))
            except (Region.DoesNotExist, ValueError):
                return Response({'error': 'Viloyat topilmadi'}, status=status.HTTP_400_BAD_REQUEST)
        elif request.user.role == 'region_admin' and request.user.region_id:
            region = request.user.region

        # Admin rollarda Django'ning is_staff=True bo'lishi kerak —
        # bu Django admin paneliga kirishga va is_staff'ga tayanadigan
        # permissionlarga ham ruxsat beradi.
        is_staff_flag = role in {'super_admin', 'region_admin', 'staff'}

        u = User.objects.create(
            phone=phone,
            username=phone,
            full_name=request.data.get('full_name') or '',
            email=request.data.get('email') or None,
            role=role,
            region=region,
            workplace=request.data.get('workplace') or '',
            job_title=request.data.get('job_title') or '',
            is_active=True,
            is_staff=is_staff_flag,
            is_onboarded=False,
        )
        # password — agar berilgan bo'lsa, set qil; aks holda usable_password emas
        password = request.data.get('password')
        if password:
            u.set_password(password)
            u.save(update_fields=['password'])
        else:
            u.set_unusable_password()
            u.save(update_fields=['password'])

        _write_audit(request.user, 'create_user', u.id, {'phone': phone, 'role': role})
        return Response(_serialize_user(u, request=request), status=status.HTTP_201_CREATED)


# ════════════════════════════════════════════════════════════════════
# RESET PASSWORD
# ════════════════════════════════════════════════════════════════════
class AdminUserResetPasswordView(APIView):
    permission_classes = [AdminUserPermission]

    def post(self, request, user_id):
        if request.user.role not in WRITE_ROLES:
            return Response({'error': 'Ruxsat yo\'q'}, status=status.HTTP_403_FORBIDDEN)
        try:
            u = _scope_queryset(request, User.objects.filter(deleted_at__isnull=True)).get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'Foydalanuvchi topilmadi'}, status=status.HTTP_404_NOT_FOUND)
        new_password = request.data.get('password')
        if not new_password or len(new_password) < 6:
            return Response({'error': 'Parol kamida 6 ta belgi'}, status=status.HTTP_400_BAD_REQUEST)
        u.set_password(new_password)
        u.save(update_fields=['password'])
        _write_audit(request.user, 'reset_user_password', u.id, {})
        return Response({'id': str(u.id), 'success': True})


# ════════════════════════════════════════════════════════════════════
# REGIONS LIST
# ════════════════════════════════════════════════════════════════════
class AdminRegionListView(APIView):
    permission_classes = [AdminUserPermission]

    def get(self, request):
        regions = Region.objects.all().order_by('name_uz')
        return Response({
            'results': [
                {
                    'id': r.id,
                    'name_uz': r.name_uz,
                    'name_ru': r.name_ru or r.name_uz,
                    'code': r.code,
                }
                for r in regions
            ]
        })


# ════════════════════════════════════════════════════════════════════
# EXPORT (Excel)
# ════════════════════════════════════════════════════════════════════
class AdminUserExportView(APIView):
    permission_classes = [AdminUserPermission]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl kerak'}, status=503)

        qs = _scope_queryset(
            request,
            User.objects.filter(deleted_at__isnull=True).select_related('region')
        )

        role = request.query_params.get('role')
        if role:
            qs = qs.filter(role=role)
        is_active = request.query_params.get('is_active')
        if is_active in ('true', '1'):
            qs = qs.filter(is_active=True)
        elif is_active in ('false', '0'):
            qs = qs.filter(is_active=False)

        qs = qs.order_by('-created_at')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Foydalanuvchilar'

        headers = ['№', 'F.I.O', 'Telefon', 'Email', 'Rol', 'Hudud', 'Ish joyi',
                   'Lavozim', 'Holat', 'Ro\'yxatdan o\'tgan', 'Oxirgi kirish']
        ws.append(headers)
        header_fill = PatternFill(start_color='1A56A0', end_color='1A56A0', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for i, u in enumerate(qs, start=1):
            ws.append([
                i,
                u.full_name or '',
                u.phone,
                u.email or '',
                dict(User.ROLE_CHOICES).get(u.role, u.role),
                u.region.name_uz if u.region else '',
                u.workplace or '',
                u.job_title or '',
                'Faol' if u.is_active else 'Faol emas',
                u.created_at.strftime('%Y-%m-%d %H:%M') if u.created_at else '',
                u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else '',
            ])

        # Column widths
        widths = [5, 30, 18, 28, 18, 20, 30, 22, 12, 18, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"uff_foydalanuvchilar_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
