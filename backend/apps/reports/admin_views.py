"""
Admin Reports API — to'liq boshqaruv: statistika, ro'yxat, generatsiya, yuklab olish, o'chirish.
"""
import os
import io
from datetime import timedelta
from pathlib import Path

from django.conf import settings
from django.db.models import Q, Count
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Report

ADMIN_ROLES = {'super_admin', 'region_admin', 'staff', 'viewer'}
WRITE_ROLES = {'super_admin', 'region_admin', 'staff'}
DELETE_ROLES = {'super_admin'}


class AdminReportPermission(IsAuthenticated):
    def has_permission(self, request, view):
        if not super().has_permission(request, view):
            return False
        return getattr(request.user, 'role', None) in ADMIN_ROLES


def _serialize_report(r, request=None):
    file_url = ''
    if r.file_url:
        if r.file_url.startswith('http'):
            file_url = r.file_url
        elif request is not None:
            file_url = request.build_absolute_uri(r.file_url)
        else:
            file_url = r.file_url
    return {
        'id': r.id,
        'type': r.type,
        'type_display': r.get_type_display(),
        'title': r.title,
        'description': r.description or '',
        'status': r.status,
        'status_display': r.get_status_display(),
        'file_url': file_url,
        'download_url': f'/api/reports/admin/{r.id}/download/' if r.file_url else '',
        'parameters': r.parameters or {},
        'generated_by_id': str(r.generated_by_id) if r.generated_by_id else None,
        'generated_by_name': (r.generated_by.full_name or r.generated_by.phone) if r.generated_by else '',
        'generated_at': r.generated_at.isoformat() if r.generated_at else None,
        'created_at': r.created_at.isoformat() if r.created_at else None,
        'download_count': (r.parameters or {}).get('downloads', 0),
    }


# ═══════════════════════════════════════════════════════════════════
# TEMPLATES
# ═══════════════════════════════════════════════════════════════════
REPORT_TEMPLATES = [
    {
        'key': 'applications_summary',
        'title_uz': 'Arizalar bo\'yicha hisobot',
        'title_ru': 'Отчёт по заявкам',
        'description_uz': 'Belgilangan davr uchun barcha arizalar tahlili',
        'description_ru': 'Анализ всех заявок за указанный период',
        'icon': 'FileText',
        'color': '#3498DB',
    },
    {
        'key': 'licenses_summary',
        'title_uz': 'Litsenziyalar bo\'yicha hisobot',
        'title_ru': 'Отчёт по лицензиям',
        'description_uz': 'Berilgan litsenziyalar va ularning holatlari',
        'description_ru': 'Выданные лицензии и их статусы',
        'icon': 'Award',
        'color': '#27AE60',
    },
    {
        'key': 'user_activity',
        'title_uz': 'Foydalanuvchilar faoliyati',
        'title_ru': 'Активность пользователей',
        'description_uz': 'Ro\'yxatdan o\'tganlar, faollar va statistikalar',
        'description_ru': 'Регистрации, активные пользователи и статистика',
        'icon': 'Users',
        'color': '#9B59B6',
    },
    {
        'key': 'regional',
        'title_uz': 'Hududiy hisobot',
        'title_ru': 'Региональный отчёт',
        'description_uz': 'Hududlar bo\'yicha taqsimot va statistika',
        'description_ru': 'Распределение и статистика по регионам',
        'icon': 'MapPin',
        'color': '#E67E22',
    },
    {
        'key': 'monthly_activity',
        'title_uz': 'Oylik faoliyat',
        'title_ru': 'Ежемесячная активность',
        'description_uz': 'Oylik ariza va litsenziya statistikasi',
        'description_ru': 'Ежемесячная статистика заявок и лицензий',
        'icon': 'Calendar',
        'color': '#F39C12',
    },
]
TEMPLATE_KEYS = {t['key'] for t in REPORT_TEMPLATES}


# ═══════════════════════════════════════════════════════════════════
# Excel generators
# ═══════════════════════════════════════════════════════════════════
def _styled_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    return wb, Font, PatternFill, Alignment, Border, Side


def _apply_header_style(ws, row, ncols, Font, PatternFill, Alignment):
    fill = PatternFill(start_color='1A56A0', end_color='1A56A0', fill_type='solid')
    font = Font(color='FFFFFF', bold=True, size=11)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _autosize(ws, max_width=40):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), max_width)


def _parse_range(params: dict):
    """Returns (start, end) datetimes, with sensible defaults (last 30 days)."""
    now = timezone.now()
    end_str = (params or {}).get('end_date')
    start_str = (params or {}).get('start_date')
    try:
        end = timezone.datetime.fromisoformat(end_str).replace(tzinfo=now.tzinfo) if end_str else now
    except Exception:
        end = now
    try:
        start = timezone.datetime.fromisoformat(start_str).replace(tzinfo=now.tzinfo) if start_str else end - timedelta(days=30)
    except Exception:
        start = end - timedelta(days=30)
    return start, end


def _generate_applications_summary(report: Report):
    from apps.applications.models import Application
    wb, Font, PatternFill, Alignment, Border, Side = _styled_workbook()
    ws = wb.active
    ws.title = 'Arizalar'

    start, end = _parse_range(report.parameters)
    qs = Application.objects.filter(submitted_at__gte=start, submitted_at__lte=end)\
        .select_related('user', 'license_type', 'region')

    # Sheet 1: Summary
    ws.append(['Hisobot davri:', f'{start.strftime("%Y-%m-%d")} — {end.strftime("%Y-%m-%d")}'])
    ws.append(['Jami arizalar:', qs.count()])
    ws.append([])

    headers = ['#', 'Ariza ID', 'F.I.O', 'Telefon', 'Litsenziya turi', 'Hudud', 'Holat', 'Yuborilgan']
    ws.append(headers)
    _apply_header_style(ws, ws.max_row, len(headers), Font, PatternFill, Alignment)
    for i, a in enumerate(qs.order_by('-submitted_at'), 1):
        ws.append([
            i,
            str(a.id)[:8].upper(),
            (a.full_name or (a.user.full_name if a.user else '') or ''),
            a.phone or (a.user.phone if a.user else ''),
            a.license_type.name_uz if a.license_type else '',
            a.region.name_uz if a.region else '',
            a.get_status_display(),
            a.submitted_at.strftime('%Y-%m-%d %H:%M') if a.submitted_at else '',
        ])

    # Sheet 2: By status
    ws2 = wb.create_sheet('Statuslar')
    ws2.append(['Holat', 'Soni'])
    _apply_header_style(ws2, 1, 2, Font, PatternFill, Alignment)
    by_status = qs.values('status').annotate(c=Count('id')).order_by('-c')
    status_map = dict(Application.STATUS_CHOICES)
    for row in by_status:
        ws2.append([status_map.get(row['status'], row['status']), row['c']])

    # Sheet 3: By type
    ws3 = wb.create_sheet('Litsenziya turlari')
    ws3.append(['Litsenziya turi', 'Soni'])
    _apply_header_style(ws3, 1, 2, Font, PatternFill, Alignment)
    by_type = qs.filter(license_type__isnull=False).values('license_type__name_uz').annotate(c=Count('id')).order_by('-c')
    for row in by_type:
        ws3.append([row['license_type__name_uz'] or '—', row['c']])

    for sh in (ws, ws2, ws3):
        _autosize(sh)
    return wb


def _generate_licenses_summary(report: Report):
    from apps.licenses.models import License
    wb, Font, PatternFill, Alignment, _, _ = _styled_workbook()
    ws = wb.active
    ws.title = 'Litsenziyalar'

    start, end = _parse_range(report.parameters)
    qs = License.objects.filter(issued_at__gte=start, issued_at__lte=end)\
        .select_related('user', 'license_type', 'region')

    ws.append(['Hisobot davri:', f'{start.strftime("%Y-%m-%d")} — {end.strftime("%Y-%m-%d")}'])
    ws.append(['Jami litsenziyalar:', qs.count()])
    ws.append([])

    headers = ['#', 'Raqam', 'Egasi', 'Telefon', 'Litsenziya turi', 'Hudud',
               'Holat', 'Berilgan', 'Tugaydi']
    ws.append(headers)
    _apply_header_style(ws, ws.max_row, len(headers), Font, PatternFill, Alignment)
    for i, l in enumerate(qs.order_by('-issued_at'), 1):
        ws.append([
            i,
            l.license_number,
            l.user.full_name if l.user else '',
            l.user.phone if l.user else '',
            l.license_type.name_uz if l.license_type else '',
            l.region.name_uz if l.region else (l.user.region.name_uz if l.user and l.user.region else ''),
            l.computed_status,
            l.issued_at.strftime('%Y-%m-%d') if l.issued_at else '',
            l.expires_at.strftime('%Y-%m-%d') if l.expires_at else '',
        ])

    ws2 = wb.create_sheet('Holatlar')
    ws2.append(['Holat', 'Soni'])
    _apply_header_style(ws2, 1, 2, Font, PatternFill, Alignment)
    statuses = {}
    for l in qs:
        s = l.computed_status
        statuses[s] = statuses.get(s, 0) + 1
    for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
        ws2.append([s, c])

    for sh in (ws, ws2):
        _autosize(sh)
    return wb


def _generate_user_activity(report: Report):
    from django.contrib.auth import get_user_model
    from apps.applications.models import Application
    from apps.licenses.models import License
    User = get_user_model()

    wb, Font, PatternFill, Alignment, _, _ = _styled_workbook()
    ws = wb.active
    ws.title = 'Foydalanuvchilar'

    start, end = _parse_range(report.parameters)
    qs = User.objects.filter(deleted_at__isnull=True)
    new_users = qs.filter(created_at__gte=start, created_at__lte=end)

    ws.append(['Hisobot davri:', f'{start.strftime("%Y-%m-%d")} — {end.strftime("%Y-%m-%d")}'])
    ws.append(['Jami foydalanuvchilar:', qs.count()])
    ws.append(['Yangi foydalanuvchilar:', new_users.count()])
    ws.append(['Faol foydalanuvchilar:', qs.filter(is_active=True).count()])
    ws.append([])

    headers = ['#', 'F.I.O', 'Telefon', 'Rol', 'Hudud', 'Arizalar', 'Litsenziyalar',
               'Holat', 'Ro\'yxatdan o\'tgan']
    ws.append(headers)
    _apply_header_style(ws, ws.max_row, len(headers), Font, PatternFill, Alignment)

    role_map = dict(User.ROLE_CHOICES)
    for i, u in enumerate(new_users.select_related('region').order_by('-created_at'), 1):
        ws.append([
            i,
            u.full_name or '',
            u.phone,
            role_map.get(u.role, u.role),
            u.region.name_uz if u.region else '',
            Application.objects.filter(user=u).count(),
            License.objects.filter(user=u).count(),
            'Faol' if u.is_active else 'Faol emas',
            u.created_at.strftime('%Y-%m-%d') if u.created_at else '',
        ])

    ws2 = wb.create_sheet('Rollar')
    ws2.append(['Rol', 'Soni'])
    _apply_header_style(ws2, 1, 2, Font, PatternFill, Alignment)
    for row in qs.values('role').annotate(c=Count('id')).order_by('-c'):
        ws2.append([role_map.get(row['role'], row['role']), row['c']])

    for sh in (ws, ws2):
        _autosize(sh)
    return wb


def _generate_regional(report: Report):
    from apps.users.models import Region
    from django.contrib.auth import get_user_model
    from apps.applications.models import Application
    from apps.licenses.models import License
    User = get_user_model()

    wb, Font, PatternFill, Alignment, _, _ = _styled_workbook()
    ws = wb.active
    ws.title = 'Hududlar'

    ws.append(['Hisobot:', 'Hududlar bo\'yicha tahlil'])
    ws.append(['Yaratilgan:', timezone.now().strftime('%Y-%m-%d %H:%M')])
    ws.append([])

    headers = ['#', 'Hudud', 'Foydalanuvchilar', 'Murabbiylar', 'Arizalar', 'Litsenziyalar', 'Faol litsenziyalar']
    ws.append(headers)
    _apply_header_style(ws, ws.max_row, len(headers), Font, PatternFill, Alignment)

    for i, r in enumerate(Region.objects.all().order_by('name_uz'), 1):
        users_qs = User.objects.filter(region=r, deleted_at__isnull=True)
        ws.append([
            i,
            r.name_uz,
            users_qs.count(),
            users_qs.filter(role='coach').count(),
            Application.objects.filter(user__region=r).count(),
            License.objects.filter(user__region=r).count(),
            License.objects.filter(user__region=r, is_active=True,
                                  expires_at__gt=timezone.now()).count(),
        ])

    _autosize(ws)
    return wb


def _generate_monthly_activity(report: Report):
    from apps.applications.models import Application
    from apps.licenses.models import License
    from django.db.models.functions import TruncMonth

    wb, Font, PatternFill, Alignment, _, _ = _styled_workbook()
    ws = wb.active
    ws.title = 'Oylik faoliyat'

    months_back = int((report.parameters or {}).get('months', 12))
    months_back = max(1, min(months_back, 36))
    end = timezone.now()
    start = (end - timedelta(days=30 * months_back)).replace(day=1)

    ws.append(['Davr:', f'{start.strftime("%Y-%m")} — {end.strftime("%Y-%m")}'])
    ws.append([])

    apps_by_month = (
        Application.objects.filter(submitted_at__gte=start)
        .annotate(month=TruncMonth('submitted_at'))
        .values('month').annotate(c=Count('id')).order_by('month')
    )
    lic_by_month = (
        License.objects.filter(issued_at__gte=start)
        .annotate(month=TruncMonth('issued_at'))
        .values('month').annotate(c=Count('id')).order_by('month')
    )

    apps_map = {r['month'].strftime('%Y-%m') if r['month'] else '': r['c'] for r in apps_by_month}
    lic_map = {r['month'].strftime('%Y-%m') if r['month'] else '': r['c'] for r in lic_by_month}

    headers = ['Oy', 'Yangi arizalar', 'Berilgan litsenziyalar']
    ws.append(headers)
    _apply_header_style(ws, ws.max_row, len(headers), Font, PatternFill, Alignment)

    months_set = sorted(set(list(apps_map.keys()) + list(lic_map.keys())))
    for m in months_set:
        ws.append([m, apps_map.get(m, 0), lic_map.get(m, 0)])

    _autosize(ws)
    return wb


GENERATORS = {
    'applications_summary': _generate_applications_summary,
    'licenses_summary': _generate_licenses_summary,
    'user_activity': _generate_user_activity,
    'regional': _generate_regional,
    'monthly_activity': _generate_monthly_activity,
}


# ═══════════════════════════════════════════════════════════════════
# STATS
# ═══════════════════════════════════════════════════════════════════
class AdminReportStatsView(APIView):
    permission_classes = [AdminReportPermission]

    def get(self, request):
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

        qs = Report.objects.all()
        total = qs.count()
        this_month = qs.filter(created_at__gte=month_start).count()
        today = qs.filter(created_at__gte=today_start).count()
        completed = qs.filter(status='completed').count()
        failed = qs.filter(status='failed').count()

        # Aggregate download counts from parameters
        total_downloads = 0
        for r in qs.only('parameters'):
            total_downloads += (r.parameters or {}).get('downloads', 0) or 0

        by_type = list(qs.values('type').annotate(count=Count('id')).order_by('-count'))
        type_map = dict(Report.TYPE_CHOICES)
        for row in by_type:
            row['type_display'] = type_map.get(row['type'], row['type'])

        return Response({
            'total': total,
            'this_month': this_month,
            'today': today,
            'completed': completed,
            'failed': failed,
            'downloads': total_downloads,
            'by_type': by_type,
        })


# ═══════════════════════════════════════════════════════════════════
# TEMPLATES LIST
# ═══════════════════════════════════════════════════════════════════
class AdminReportTemplatesView(APIView):
    permission_classes = [AdminReportPermission]

    def get(self, request):
        return Response({'results': REPORT_TEMPLATES})


# ═══════════════════════════════════════════════════════════════════
# LIST
# ═══════════════════════════════════════════════════════════════════
class AdminReportListView(APIView):
    permission_classes = [AdminReportPermission]

    def get(self, request):
        qs = Report.objects.select_related('generated_by').order_by('-created_at')

        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(title__icontains=search) | Q(description__icontains=search))

        rtype = (request.query_params.get('type') or '').strip()
        if rtype:
            qs = qs.filter(type=rtype)

        rstatus = (request.query_params.get('status') or '').strip()
        if rstatus:
            qs = qs.filter(status=rstatus)

        try:
            limit = max(1, min(int(request.query_params.get('limit', 20)), 200))
            offset = max(0, int(request.query_params.get('offset', 0)))
        except ValueError:
            limit, offset = 20, 0

        count = qs.count()
        page = qs[offset:offset + limit]
        return Response({
            'count': count, 'limit': limit, 'offset': offset,
            'results': [_serialize_report(r, request=request) for r in page],
        })


# ═══════════════════════════════════════════════════════════════════
# DETAIL
# ═══════════════════════════════════════════════════════════════════
class AdminReportDetailView(APIView):
    permission_classes = [AdminReportPermission]

    def get(self, request, report_id):
        try:
            r = Report.objects.select_related('generated_by').get(id=report_id)
        except Report.DoesNotExist:
            return Response({'error': 'Hisobot topilmadi'}, status=status.HTTP_404_NOT_FOUND)
        return Response(_serialize_report(r, request=request))


# ═══════════════════════════════════════════════════════════════════
# GENERATE
# ═══════════════════════════════════════════════════════════════════
class AdminReportGenerateView(APIView):
    permission_classes = [AdminReportPermission]

    def post(self, request):
        if request.user.role not in WRITE_ROLES:
            return Response({'error': 'Ruxsat yo\'q'}, status=status.HTTP_403_FORBIDDEN)

        template = (request.data.get('template') or '').strip()
        if template not in TEMPLATE_KEYS:
            return Response(
                {'error': 'Noto\'g\'ri shablon', 'valid': list(TEMPLATE_KEYS)},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Determine type bucket from template
        type_map = {
            'monthly_activity': 'monthly',
            'applications_summary': 'custom',
            'licenses_summary': 'custom',
            'user_activity': 'custom',
            'regional': 'custom',
        }
        rtype = type_map.get(template, 'custom')

        # Resolve template metadata
        meta = next((t for t in REPORT_TEMPLATES if t['key'] == template), None)
        custom_title = (request.data.get('title') or '').strip()

        # Til: payload'dan yoki Accept-Language header'dan aniqlaymiz
        lang = (request.data.get('language') or '').lower().strip()
        if lang not in ('uz', 'ru'):
            accept = (request.headers.get('Accept-Language') or '').lower()
            lang = 'ru' if accept.startswith('ru') else 'uz'

        title_key = 'title_ru' if lang == 'ru' else 'title_uz'
        desc_key = 'description_ru' if lang == 'ru' else 'description_uz'
        title = custom_title or (meta[title_key] if meta else template)

        params = {
            'template': template,
            'start_date': request.data.get('start_date'),
            'end_date': request.data.get('end_date'),
            'months': request.data.get('months'),
            'language': lang,
            'downloads': 0,
        }

        report = Report.objects.create(
            type=rtype,
            title=title,
            description=meta[desc_key] if meta else '',
            status='generating',
            parameters=params,
            generated_by=request.user,
        )

        # Generate Excel
        try:
            try:
                from openpyxl import Workbook  # noqa
            except ImportError:
                report.status = 'failed'
                report.save()
                return Response({'error': 'openpyxl o\'rnatilmagan'}, status=503)

            generator = GENERATORS[template]
            wb = generator(report)

            # Save to media/reports/
            reports_dir = Path(settings.MEDIA_ROOT) / 'reports'
            reports_dir.mkdir(parents=True, exist_ok=True)
            filename = f"report_{report.id}_{template}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = reports_dir / filename
            wb.save(str(filepath))

            report.file_url = f'{settings.MEDIA_URL}reports/{filename}'
            report.status = 'completed'
            report.generated_at = timezone.now()
            report.save()
        except Exception as e:
            report.status = 'failed'
            report.description = f'{report.description} | XATO: {str(e)[:200]}'
            report.save()
            return Response(
                {'error': f'Generatsiya xatoligi: {str(e)}', 'report': _serialize_report(report, request=request)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(_serialize_report(report, request=request), status=status.HTTP_201_CREATED)


# ═══════════════════════════════════════════════════════════════════
# DOWNLOAD
# ═══════════════════════════════════════════════════════════════════
class AdminReportDownloadView(APIView):
    permission_classes = [AdminReportPermission]

    def get(self, request, report_id):
        try:
            r = Report.objects.get(id=report_id)
        except Report.DoesNotExist:
            return Response({'error': 'Hisobot topilmadi'}, status=status.HTTP_404_NOT_FOUND)
        if not r.file_url or r.status != 'completed':
            return Response({'error': 'Fayl mavjud emas'}, status=status.HTTP_404_NOT_FOUND)

        rel_path = r.file_url
        if rel_path.startswith(settings.MEDIA_URL):
            rel_path = rel_path[len(settings.MEDIA_URL):]
        full_path = os.path.join(settings.MEDIA_ROOT, rel_path)
        if not os.path.exists(full_path):
            return Response({'error': 'Fayl topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        # Bump download count
        params = r.parameters or {}
        params['downloads'] = int(params.get('downloads', 0)) + 1
        r.parameters = params
        r.save(update_fields=['parameters'])

        response = FileResponse(open(full_path, 'rb'),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = os.path.basename(full_path)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ═══════════════════════════════════════════════════════════════════
# DELETE
# ═══════════════════════════════════════════════════════════════════
class AdminReportDeleteView(APIView):
    permission_classes = [AdminReportPermission]

    def delete(self, request, report_id):
        if request.user.role not in DELETE_ROLES:
            return Response({'error': 'Faqat super admin o\'chira oladi'}, status=status.HTTP_403_FORBIDDEN)
        try:
            r = Report.objects.get(id=report_id)
        except Report.DoesNotExist:
            return Response({'error': 'Hisobot topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        # Remove file
        if r.file_url:
            rel = r.file_url[len(settings.MEDIA_URL):] if r.file_url.startswith(settings.MEDIA_URL) else r.file_url
            full = os.path.join(settings.MEDIA_ROOT, rel)
            try:
                if os.path.exists(full):
                    os.remove(full)
            except Exception:
                pass

        r.delete()
        return Response({'id': report_id, 'deleted': True})
